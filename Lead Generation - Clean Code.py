#!python 3
import pandas as pd 
import numpy as np 
import glob
import time
from tkinter import *
from tkinter import scrolledtext, messagebox, ttk, filedialog
from tkinter.ttk import Progressbar
from openpyxl import load_workbook 
from src.funAppender import appender
from src.cepChecker import requestcodeValidation, checkCEP

all_data = pd.DataFrame()

def selectFile():
       excelFile = filedialog.askopenfilename(filetypes = (("Excel file", "*.xlsx"), ("all files", "*.*")))
       return excelFile

def regex(yourDataFrame):
       yourDataFrame.replace(to_replace ="^(\+55|55|055|\s\+55|\s55|\s055){0,1}?(\s|-|\.|\+|\_){0,2}?(0{0,1}\s{0,1})?([1-9][0-9]|\([1-9][0-9]\)){0,1}(\s|-|\.|\+|\_){0,2}?((?!9999)\d{4,5}){0,1}?(\s|-|\.|\+|\_)?((?!0000)\d{4,5})$", value=r"\4-\6\8", regex=True, inplace=True)
       yourDataFrame.filter(regex="^(\+55|55|055|\s\+55|\s55|\s055){0,1}?(\s|-|\.|\+|\_){0,2}?(0{0,1}\s{0,1})?([1-9][0-9]|\([1-9][0-9]\)){0,1}(\s|-|\.|\+|\_){0,2}?(\d{4,5}){0,1}?(\s|-|\.|\+|\_)?((?!0000)\d{4,5})$", axis=1)
       return yourDataFrame

def dropRegex(yourDataFrame, length):
       yourDataFrame = yourDataFrame.drop(yourDataFrame[yourDataFrame['Phone Mask'].map(len) < length ].index)
       return yourDataFrame

def saveFile (yourDataFrame):
       yourDataFrame = yourDataFrame.to_excel(writer, sheet_name='Base Original') #Excel sheet which it'll be pasted on
       writer.save()
       return yourDataFrame
       
def readAsDataframe(excelFile):
       df = pd.read_excel(excelFile)
       templateBook = excelFile 
       global writer
       book = load_workbook(templateBook)
       writer = pd.ExcelWriter(templateBook, engine='openpyxl')
       writer.book = book
       writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
       return df

def rowIteration(df):
       df.insert(19, column='verificacao', value = 'none')
       for index, row in df.iterrows():
              zxzx = checkCEP(str(row["Primary Postcode"]))
              df['verificacao'][index] = zxzx
              time.sleep(15)
       print('rowIteration concluded')
       print(df.head())
       return df


### COMMMAND BUTTONS ###

def excelAccess():
       templateBook = selectFile()
       global writer
       book = load_workbook(templateBook)
       writer = pd.ExcelWriter(templateBook, engine='openpyxl')
       writer.book = book
       writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

def runButton():
       saveFile(dropRegex(regex(appender(all_data)), 11))
       print('Finalizado!')

def cepButton():
       saveFile(rowIteration(readAsDataframe(selectFile())))
       print('Finalizado!')
       


##################### 

######################### - USER INTERFACE  - ##########################

window = Tk()
window.title("GPBR-DBM")
window.geometry('180x150')

desc = Label(window, text="Selecione seu arquivo de destino\n e qual função a ser utilizada")
desc.grid(column=0, row=0)

selectFun = Label(window, text="Selecione sua função")
selectFun.grid(column=0, row=4)

templateButton = Button(window, text="Template", command=excelAccess)
templateButton.grid(column=0, row=3)

runButton = Button(window, text="Concatenar listas Hubspot", command=runButton)
runButton.grid(column=0, row=5)

runcepButton = Button(window, text="Checagem de CEP", command=cepButton)
runcepButton.grid(column=0, row=6)

window.mainloop()

########################################################################