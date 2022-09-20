print ("Importing Modules")
import pandas as pd
import numpy as np
import glob
import re
from openpyxl import Workbook, load_workbook
import json
from simple_salesforce import Salesforce

pd.options.mode.chained_assignment = None  # default='warn'
j = pd.DataFrame(columns=['List','Count'])
print("Analyzing files...")


campaign_list_ref = {
"":""
}

all_data = pd.DataFrame()
for files in glob.glob('*.xlsx'):
   appealCampaign = re.search(".*", files)
   df = pd.read_excel(files, dtype={'First Name':'string','Last Name':'string', 'Email':'string', 'Phone Number':'string', 'Phone Mask':'string'}).assign(Last_Appeal_Name__c=appealCampaign.group()[:-5])
   df.insert(6, column= 'CampaignID', value = campaign_list_ref[df.iloc[0]['Last_Appeal_Name__c']])
   all_data = all_data.append(df, ignore_index=True)
   j.loc[files] = [[appealCampaign.group()[:-5]] , len(df.index)]
   print (all_data.shape[0])
   
print("All files have been concatenated. Accessing folder...")

file = Workbook().save(filename="Template.xlsx")
book = load_workbook('Template.xlsx')
writer = pd.ExcelWriter('Template.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
print("Running Regex")

all_data["Phone Number"].replace(to_replace="^(\+55|55|055|\s\+55|\s55|\s055){0,1}?(\s|-|\.|\+|\_)?(0{0,1}\s{0,1})?([1-9][0-9]|\([1-9][0-9]\))(\s|-|\.|\+|\_|\(|\)){0,3}?((?!9999)\d{4,5}){0,1}?(\s|-|\.|\+|\_)?((?!0000)?\d{4,5})$", value=r"\4-\6\8", regex=True, inplace=True)
all_data["Phone Mask"].replace(to_replace="^(\+55|55|055|\s\+55|\s55|\s055){0,1}?(\s|-|\.|\+|\_)?(0{0,1}\s{0,1})?([1-9][0-9]|\([1-9][0-9]\))(\s|-|\.|\+|\_|\(|\)){0,3}?((?!9999)\d{4,5}){0,1}?(\s|-|\.|\+|\_)?((?!0000)?\d{4,5})$", value=r"\4-\6\8", regex=True, inplace=True)
#all_data.replace(to_replace ={'Phone Number':"^(\+55|55|055|\s\+55|\s55|\s055){0,1}?(\s|-|\.|\+|\_)?(0{0,1}\s{0,1})?([1-9][0-9]|\([1-9][0-9]\))(\s|-|\.|\+|\_|\(|\)){0,3}?((?!9999)\d{4,5}){0,1}?(\s|-|\.|\+|\_)?((?!0000)?\d{4,5})$"}, value=r"\4-\6\8", regex=True, inplace=True)

### As funções de Replace e Filter no dataframe aplicam a Regex inconsistentemente em relação ao regex101.com
### Possivelmente vale o teste de copiar a coluna de telefone em um dataframe próprio, aplicar o regex, e então retorná-lo pro DF original
### Alternativamente, talvez valha a pena aplicar esse filtro célula a célula.
print("Regex sucessfully applied. Excluding unmatched results")

all_data.filter(regex="^(\+55|55|055|\s\+55|\s55|\s055){0,1}?(\s|-|\.|\+|\_)?(0{0,1}\s{0,1})?([1-9][0-9]|\([1-9][0-9]\))(\s|-|\.|\+|\_){0,2}?(\d{4,5}){0,1}?(\s|-|\.|\+|\_)?((?!0000)?\d{4,5})$", axis=1)

all_data.insert(7, column='validacao de len [Phone Number]', value = 'No issues')
all_data.insert(8, column='validacao de len [Phone Mask]', value = 'No issues')
all_data.insert(9, column='Lead Source', value = 'Organic')
all_data.insert(10, column='Partner', value = 'None')
all_data.insert(11,column='company', value = '.')

paid_utm_sources = ["paid", "cpc", "partner", "adc", "ad", "ppc", "cpm", "cpv", "adglow" ] # Lists all paid tags from utm_source

print('Validating Phone Number Column...')

for index, row  in all_data.iterrows():
   ad = all_data['Phone Number'][index]
   try:
      if len(ad) == 12 and ad[3] != '9':
         all_data['validacao de len [Phone Number]'][index] = 'Não começa com 9'
      elif ad[1] == '0':
         all_data['validacao de len [Phone Number]'][index] = 'DDD inválido'
      elif ad[0] == '0':
         all_data['validacao de len [Phone Number]'][index] = 'DDD inválido'
      elif len(ad) < 10:
         all_data['validacao de len [Phone Number]'][index] = 'Curto demais'
      elif len(ad) > 13:
         all_data['validacao de len [Phone Number]'][index] = 'Longo demais'
      elif len(ad) == 11 and (ad[3] == '6' or ad[3] == '7' or ad[3] == '8' or ad[3] == '9'):
         all_data['validacao de len [Phone Number]'][index] = 'Não tem o nono dígito'
      elif len(ad) == 'NaN':
         all_data['validacao de len [Phone Number]'][index] = 'Telefone não informado'          
   except:
      all_data['validacao de len [Phone Number]'][index] = 'Não conseguiu identificar o número'
         
# print('Validating Phone Number Region...')
# for index, row  in all_data.iterrows():
#    ad = all_data['Phone Number'][index]
#    try:
#       if ad[0] == '7' and (ad[1] == '1'or'3'or'4'or'5'or'7'or'9'):
#          all_data['validacao de len [Phone Number]'][index] = 'Região filtrada'
#       elif ad[0] == '8' and (ad[1] == '1'or'2'or'3'or'4'or'5'or'6'or'7'or'8'or'9'):
#          all_data['validacao de len [Phone Number]'][index] = 'Região filtrada'
#       elif ad[0] == '9' and (ad[1] == '8'or'9'):
#          all_data['validacao de len [Phone Number]'][index] = 'Região filtrada'
#    except:
#      all_data['validacao de len [Phone Number]'][index] = 'Não conseguiu identificar o número'

print('Validating Phone Mask Column...')
for index, row  in all_data.iterrows():
   ad = all_data['Phone Mask'][index]
   try:
      if len(ad) == 12 and ad[3] != '9':
         all_data['validacao de len [Phone Mask]'][index] = 'Não começa com 9'
      elif ad[1] == '0':
         all_data['validacao de len [Phone Mask]'][index] = 'DDD inválido'
      elif ad[0] == '0':
         all_data['validacao de len [Phone Mask]'][index] = 'DDD inválido'
      elif len(ad) < 10:
         all_data['validacao de len [Phone Mask]'][index] = 'Curto demais'
      elif len(ad) > 12:
         all_data['validacao de len [Phone Mask]'][index] = 'Longo demais'
      #elif len(ad) == 11 and (ad[3] == '6' or ad[3] == '7' or ad[3] == '8' or ad[3] == '9'):
      #   all_data['validacao de len [Phone Mask]'][index] = 'Não tem o nono dígito'
      elif len(ad) == 'NaN':
         all_data['validacao de len [Phone Mask]'][index] = 'Telefone não informado'
   except:
      if all_data['validacao de len [Phone Number]'][index] == 'Não conseguiu identificar o número':
         all_data['validacao de len [Phone Mask]'][index] = 'Não conseguiu identificar o número'
      else:
         all_data['validacao de len [Phone Mask]'][index] = all_data['validacao de len [Phone Number]'][index]
print (all_data.shape[0])

print('Setting partner field...')
for index, row  in all_data.iterrows():
   partner = all_data['Last_Appeal_Name__c'][index]
   try:
      if 'Munduruku' in partner:
         all_data['Partner'][index] = 'Munduruku - Organic'
      elif  'Actual - Forests - Todos' in  partner:
         all_data['Partner'][index] = 'TODOS - Actual Sales'
      elif  'Adsplay - Forests - Todos ' in  partner:
         all_data['Partner'][index] = 'TODOS - Adsplay'
      elif 'Pareto e CPC - Forests - Todos' in  partner:
         all_data['Partner'][index] = 'TODOS - Pareto'
      elif 'Lets Perf - Forests - Todos' in partner:
         all_data['Partner'][index] = "TODOS - Let's Perform"
      elif 'Petição' in partner:
         all_data['Partner'][index] = 'Organic'
      elif 'Beeleads - Forests - Todos' in partner:
         all_data['Partner'][index] = 'TODOS - Beeleads'
      elif 'Organic - Forests - Todos' in partner:
         all_data['Partner'][index] = 'TODOS - Organic'
      elif 'Army Help The Planet - Forests - Todos' in partner:
         all_data['Partner'][index] = 'TODOS - Army Help The Planet'
      elif 'Organic - Forests - Munduruku' in partner:
         all_data['Partner'][index] = 'Munduruku - Organic'
      elif 'Pareto Paid - Forest - Munduruku' in partner:
         all_data['Partner'][index] = 'Munduruku - Pareto'         
      elif 'Agroecologia' in partner:
         all_data['Partner'][index] = 'Agroecologia - Pareto'     
      elif 'Pareto Paid - Forest - Munduruku' in partner:
         all_data['Partner'][index] = 'Munduruku - Pareto'
      elif 'Adsplay - Forests - Todos' in partner:
         all_data['Partner'][index] = 'TODOS - Adsplay'  
   except:
      raise Exception

# print('Validating Phone Number Column...')
# for index, row  in all_data.iterrows():
#    ad = all_data['Create Date'][index]
#    try:
#       if ad.month == 2:
#          all_data['CampaignID'][index] = campaign_list_feb[all_data['CampaignID'][index]]
#    except:
#       raise Exception


# print('Validating Phone Mask Region...')
# for index, row  in all_data.iterrows():
#    ad = all_data['Phone Mask'][index]
#    try:
#       if ad[0] == '7' and (ad[1] == '1'or'3'or'4'or'5'or'7'or'9'):
#          all_data['validacao de len [Phone Mask]'][index] = 'Região filtrada'
#       elif ad[0] == '8' and (ad[1] == '1'or'2'or'3'or'4'or'5'or'6'or'7'or'8'or'9'):
#          all_data['validacao de len [Phone Mask]'][index] = 'Região filtrada'
#       elif ad[0] == '9' and (ad[1] == '8'or'9'):
#          all_data['validacao de len [Phone Mask]'][index] = 'Região filtrada'
#    except:
#       all_data['validacao de len [Phone Mask]'][index] = 'Não conseguiu identificar o número'

print('Creating Analysis DataFrame...')

analysisDF = all_data[['validacao de len [Phone Mask]', 'validacao de len [Phone Number]']].copy()
analysisDF = analysisDF.drop(columns = ['validacao de len [Phone Mask]', 'validacao de len [Phone Number]'])
print('Matching Organic/Paid Labels')
for index, row  in all_data.iterrows():
   source = all_data["Utm medium"][index]
   try:
      if source in paid_utm_sources:
         all_data["Lead Source"][index] = "Paid"
   except:
      raise Exception

analysisDF.insert(0, column='result', value = 'Ambos os números estavam corretos')
analysisDF.insert(1, column='source', value = 'Organic')
analysisDF.insert(2, column='partner', value = 'None')
analysisDF["source"] = all_data["Lead Source"]
analysisDF["partner"] = all_data["Partner"]

print('Excluding filtered numbers')
for index, row  in all_data.iterrows():
   vlen_pm = all_data['validacao de len [Phone Mask]'][index]
   vlen_pn = all_data['validacao de len [Phone Number]'][index]
   try:
      if vlen_pm != 'No issues' and vlen_pn != 'No issues' :
         analysisDF['result'][index] = all_data['validacao de len [Phone Mask]'][index]
         #all_data.drop(index, inplace=True)
   except:
      raise Exception
   try:
      if vlen_pm == 'No issues' and vlen_pn != 'No issues' :
         all_data['validacao de len [Phone Number]'][index] = all_data['validacao de len [Phone Mask]'][index]
         all_data['Phone Number'][index] = all_data['Phone Mask'][index]
         analysisDF['result'][index] = 'Phone Mask estava correto'
      elif vlen_pn == 'No issues' and vlen_pm != 'No issues':
         all_data['validacao de len [Phone Mask]'][index] = all_data['validacao de len [Phone Number]'][index]
         all_data['Phone Mask'][index] = all_data['Phone Number'][index]
         analysisDF['result'][index] = 'Phone Number estava correto'
      elif vlen_pm == 'Não conseguiu identificar o número' and vlen_pn != 'Não conseguiu identificar o número':
         vlen_pm = vlen_pn

   except:
      raise Exception
print (all_data.shape[0])


print("Saving your compiled list...")

analysisDF.to_excel('output.xlsx', index = False)
all_data = all_data.drop(columns=["Company ID","Company name", "validacao de len [Phone Number]","validacao de len [Phone Mask]","Utm medium","Phone Mask","Contact ID", "Lead Source","Partner", "company","Last_Appeal_Name__c"])

all_data.to_excel(writer, sheet_name='Sheet')
writer.save()
print ("Done")

print("Generating Analysis Summary Files...")

DF = pd.read_excel('output.xlsx')
fd = DF
DF = DF.drop(columns=["source"])
print(DF.head())
valid = DF.copy()
valid = valid.loc[valid['result'].isin(['Phone Mask estava correto','Phone Number estava correto','Ambos os números estavam corretos'])]
print(valid.head())
invalid = DF.copy()
invalid = invalid.loc[invalid['result'].isin(['Não tem o nono dígito','Não conseguiu identificar o número','Curto demais','DDD inválido','Longo demais','Não tem o nono dígito', 'Não começa com 9'])]
print(invalid.head())
v = valid.groupby('partner')['result']
i = invalid.groupby('partner')['result']
main = pd.concat([v.value_counts()],axis=1, keys=('count'))
main2 = pd.concat([i.value_counts()],axis=1, keys=('count'))

# main = pd.concat([g.value_counts(), 
#                 DF.value_counts(normalize=True).mul(100).round(1)],axis=0, keys=('count'))

print('Generating Summary...')

f = fd.groupby('source').count()
percen = fd.value_counts('source',  normalize=True).mul(100).round(1)
f['percentage'] = percen

print('Saving Analysis DataFrame...')

def multiple_dfs(df_list, sheets, file_name, spaces):
    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
    row = 0
    for dataframe in df_list:
        dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
        row = row + len(dataframe.index) + spaces + 1
    writer.save()

    # list of dataframes
dfs = [j, main, main2, f]
print(dfs)

# run function
multiple_dfs(dfs, 'Validation', 'test2.xlsx', 3)

print("Done")


# # function
# def dfs_tabs(df_list, sheet_list, file_name):
#     writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
#     for dataframe, sheet in zip(df_list, sheet_list):
#         dataframe.to_excel(writer, sheet_name=sheet, startrow=0 , startcol=0)   
#     writer.save()

# # list of dataframes and sheet names
# dfs = [df, df1, df2]
# sheets = ['df','df1','df2']    

# # run function
# dfs_tabs(dfs, sheets, 'multi-test.xlsx')


print("Uploading to Salesforce")
df = pd.DataFrame()

df = pd.read_excel("Template.xlsx")

sf = Salesforce (
    username = "",
    password = "",
    security_token = "",
    instance_url = ""
)
grouped = df.groupby(df['CampaignID'])
for ID in df['CampaignID'].unique():
    temporary_df = grouped.get_group(ID)
    
    appeal = sf.query(
        f"SELECT GP_Appeal_Code__c FROM Campaign WHERE ID = '{ID}'")["records"][0]["GP_Appeal_Code__c"]
    campaignName = sf.query(
        f"SELECT Name FROM Campaign WHERE ID = '{ID}'")["records"][0]["Name"]
    sf.Campaign.update(f'{ID}',{'NumberSent': f'{len(temporary_df.index)}'})
    temporary_df = temporary_df.iloc[: , 1:]
    temporary_df.to_excel(f'{appeal} - {campaignName}.xlsx', index = False)


